"""Spreadsheet importer — migrate the household workbook into the budget plan.

Reads an ``.xlsx`` household calculation and produces a
:class:`~.budget_plan.BudgetPlan`. This is the migration path: the household
keeps its years of accumulated structure instead of retyping ~30 positions.

Expected sheets (matched case-insensitively, extra sheets are ignored):

``Breakdown Einkommen``
    Column headers name the persons plus one aggregate column. Rows carry the
    income deposit, the mandatory private insurance and the tax settlement.

``Breakdown Kosten``
    Column A is the position name, one column per person plus one shared
    column holds the amount, and a trailing column may hold a comment. A
    position's amount appears in exactly one owner column.

Structure the importer recovers that a plain value read would lose:

- **Buffer positions** are written as ``=80*4.5`` (unit price x units) so the
  factors stay adjustable. Positions whose name marks them as a buffer keep
  both factors instead of being flattened to a product.
- **Validity windows** are written in prose in the comment column
  ("bis inkl. April 2027") and become a machine-readable ``valid_until``.
- **Reimbursements** are negative amounts (a shared subscription billed onward
  to a third party) and must stay negative rather than becoming income.

SECURITY: the workbook holds real financial data. It is read from a path the
user supplies, parsed in memory, and written only to HA's ``.storage/``. The
file itself is never copied into the repository and its contents are never
logged — only counts and structural warnings.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from typing import Any

from .budget_plan import BudgetPlan, CostPosition, IncomeEntry, slugify_position
from .const import (
    CATEGORIZATION_RULES,
    CATEGORY_FOOD,
    CATEGORY_HOUSING,
    CATEGORY_INSURANCE,
    CATEGORY_LOANS,
    CATEGORY_OTHER,
    CATEGORY_SUBSCRIPTIONS,
    CATEGORY_TRANSPORT,
    CATEGORY_UTILITIES,
    OWNER_SHARED,
    POSITION_KIND_BUFFER,
    POSITION_KIND_FIXED,
)

_LOGGER = logging.getLogger(__name__)

SHEET_INCOME = "breakdown einkommen"
SHEET_COSTS = "breakdown kosten"

# Header labels that mark the aggregate/shared column rather than a person.
_SHARED_HEADERS = {"gemeinsam", "shared", "beide", "joint", "total", "gesamt", "summe"}

# Header labels that mark a free-text column. Without this, a "Kommentar"
# header is taken for a person's name, the comment column is looked for one
# column further right, and every validity window in the sheet is lost.
_COMMENT_HEADERS = {
    "kommentar",
    "kommentare",
    "comment",
    "comments",
    "note",
    "notes",
    "notiz",
    "bemerkung",
    "hinweis",
}

# Row labels in the income sheet, matched as substrings (lower-case).
_ROW_DEPOSIT = ("einkommen (einzahlung", "einzahlung", "gehalt", "deposit")
_ROW_INSURANCE = ("privatversicherung", "private insurance", "versicherung")
_ROW_TAX = ("steuerausgleich", "steuerklasse", "tax")
_ROW_NET = ("netto", "(netto)", "net")

# A position whose name contains one of these is a budgeted buffer, not a
# fixed debit. Buffers are transferred and debited separately.
_BUFFER_MARKERS = ("puffer", "buffer")

# Positions that are sums of several contracts, not a single debit.
_MULTIPLY_FORMULA = re.compile(r"^=\s*([\d]+(?:[.,]\d+)?)\s*\*\s*([\d]+(?:[.,]\d+)?)\s*$")

_MONTHS_DE = {
    "januar": 1, "februar": 2, "märz": 3, "maerz": 3, "april": 4,
    "mai": 5, "juni": 6, "juli": 7, "august": 8, "september": 9,
    "oktober": 10, "november": 11, "dezember": 12,
}

# Position-name keywords that override the generic categorizer rules. The
# generic rules are tuned for bank transaction text and mis-file some position
# names — most notably a loan whose creditor is a health insurer.
_NAME_CATEGORY_OVERRIDES: list[tuple[tuple[str, ...], str]] = [
    (("kredit", "darlehen", "tilgung", "klarna"), CATEGORY_LOANS),
    (("miete", "warm", "hausrat"), CATEGORY_HOUSING),
    (("essen", "hellofresh", "haushalt einkaufen", "lebensmittel"), CATEGORY_FOOD),
    (("ticket", "bahn", "celular", "mobil"), CATEGORY_TRANSPORT),
    (("versicherung", "haftpflicht"), CATEGORY_INSURANCE),
    (
        ("strom", "wasser", "internet", "warmwasser", "rundfunk", "telekom", "vodafone"),
        CATEGORY_UTILITIES,
    ),
    (("netflix", "youtube", "google", "prime", "claude", "home assistant"), CATEGORY_SUBSCRIPTIONS),
]


@dataclass
class ImportReport:
    """Outcome of an import run — safe to log and to show in the UI.

    Contains counts and structural notes only, never amounts.
    """

    persons: list[str] = field(default_factory=list)
    positions_imported: int = 0
    shared_positions: int = 0
    buffer_positions: int = 0
    reimbursements: int = 0
    time_boxed: int = 0
    income_entries: int = 0
    warnings: list[str] = field(default_factory=list)
    net_income_check: str = ""

    def to_dict(self) -> dict[str, Any]:
        """Serialize for the service response."""
        return {
            "persons": self.persons,
            "positions_imported": self.positions_imported,
            "shared_positions": self.shared_positions,
            "buffer_positions": self.buffer_positions,
            "reimbursements": self.reimbursements,
            "time_boxed": self.time_boxed,
            "income_entries": self.income_entries,
            "warnings": self.warnings,
            "net_income_check": self.net_income_check,
        }


def categorize_position(name: str) -> str:
    """Map a position name to a category.

    Name-specific overrides are checked first because the generic keyword rules
    are written for bank transaction text: a position called "TK Kredit" hits
    the insurance rule's ``"tk "`` keyword before the loan rule and would be
    filed as insurance.
    """
    lowered = f" {name.lower()} "

    for keywords, category in _NAME_CATEGORY_OVERRIDES:
        if any(kw in lowered for kw in keywords):
            return category

    for category, keywords in CATEGORIZATION_RULES.items():
        if any(kw in lowered for kw in keywords):
            return category

    return CATEGORY_OTHER


def parse_validity(comment: str) -> tuple[str | None, str | None]:
    """Extract a validity window from a free-text comment.

    Recognises the forms the household actually writes, e.g.
    "bis inkl. April 2027" → ``(None, "2027-04")`` and
    "ab Mai 2026" → ``("2026-05", None)``.

    Returns ``(valid_from, valid_until)``; unrecognised text yields
    ``(None, None)`` so an unparsed comment never silently drops a position.
    """
    if not comment:
        return (None, None)
    text = str(comment).lower()

    def find_month_year(segment: str) -> str | None:
        for name, number in _MONTHS_DE.items():
            if name in segment:
                match = re.search(r"(20\d{2})", segment)
                if match:
                    return f"{match.group(1)}-{number:02d}"
        match = re.search(r"(20\d{2})[-/.](\d{1,2})", segment)
        if match:
            return f"{match.group(1)}-{int(match.group(2)):02d}"
        return None

    valid_from = None
    valid_until = None

    if any(marker in text for marker in ("bis", "until", "till", "ende")):
        valid_until = find_month_year(text)
    if any(marker in text for marker in ("ab ", "from ", "seit", "start")):
        valid_from = find_month_year(text)

    return (valid_from, valid_until)


def _to_float(value: Any) -> float | None:
    """Coerce a cell value to float, returning None when it is not numeric."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    # Strip both regular and non-breaking spaces: Excel thousand separators
    # are often U+00A0, which a plain space replace leaves behind.
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if not text:
        return None
    text = text.replace("€", "")
    # German decimal comma, only when it is unambiguous.
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _is_shared_header(label: str) -> bool:
    """Whether a column header marks the shared/aggregate column."""
    return label.strip().lower() in _SHARED_HEADERS


def _find_sheet(workbook: Any, wanted: str) -> Any | None:
    """Find a sheet by case-insensitive name, tolerating stray whitespace."""
    for name in workbook.sheetnames:
        if name.strip().lower() == wanted:
            return workbook[name]
    for name in workbook.sheetnames:
        if wanted in name.strip().lower():
            return workbook[name]
    return None


def _header_map(sheet: Any, max_col: int = 12) -> tuple[dict[int, str], int | None]:
    """Read the owner columns and the comment column from the header row.

    Returns ``({column_index: owner}, comment_column)`` where owner is a
    person's name or :data:`OWNER_SHARED`. The first row containing at least
    two owner labels right of column A is taken as the header.

    Free-text columns are identified by their header and excluded from the
    owner map — otherwise a "Kommentar" column is mistaken for a person.
    """
    for row_idx in range(1, 4):
        labels: dict[int, str] = {}
        comment_col: int | None = None
        for col_idx in range(2, max_col + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            label = str(value).strip()
            if not label:
                continue
            if label.lower() in _COMMENT_HEADERS:
                if comment_col is None:
                    comment_col = col_idx
                continue
            labels[col_idx] = OWNER_SHARED if _is_shared_header(label) else label
        if len(labels) >= 2:
            return labels, comment_col
    return {}, None


def parse_income_sheet(sheet: Any, report: ImportReport) -> dict[str, IncomeEntry]:
    """Parse ``Breakdown Einkommen`` into per-person income entries."""
    headers, _ = _header_map(sheet)
    if not headers:
        report.warnings.append("Income sheet: no person columns found in the header row")
        return {}

    entries: dict[str, IncomeEntry] = {}
    for owner in headers.values():
        if owner != OWNER_SHARED:
            entries[owner] = IncomeEntry(person=owner)

    stated_net: dict[str, float] = {}

    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 40)):
        label_cell = row[0].value if row else None
        if label_cell is None:
            continue
        label = str(label_cell).strip().lower()
        if not label:
            continue

        for col_idx, owner in headers.items():
            value = _to_float(sheet.cell(row=row[0].row, column=col_idx).value)
            if value is None:
                continue
            if owner == OWNER_SHARED:
                if any(marker in label for marker in _ROW_NET) and "rel" not in label:
                    stated_net["__total__"] = value
                continue
            entry = entries.setdefault(owner, IncomeEntry(person=owner))
            if any(marker in label for marker in _ROW_NET) and "rel" not in label:
                stated_net[owner] = value
            elif any(marker in label for marker in _ROW_TAX):
                entry.tax_adjustment = value
            elif any(marker in label for marker in _ROW_INSURANCE):
                # Stored as a negative deduction regardless of the sheet's sign.
                entry.insurance_mandatory = -abs(value)
            elif any(marker in label for marker in _ROW_DEPOSIT):
                entry.deposit = value

    # Cross-check our arithmetic against the sheet's own stated net income.
    # A mismatch means we mapped a row wrongly — surface it instead of
    # shipping a plan that silently disagrees with the spreadsheet.
    mismatches = []
    for person, entry in entries.items():
        if person in stated_net and abs(entry.net - stated_net[person]) > 0.02:
            mismatches.append(person)
    report.income_entries = len(entries)
    report.net_income_check = (
        "matches spreadsheet"
        if not mismatches
        else f"differs for: {', '.join(sorted(mismatches))}"
    )
    if mismatches:
        report.warnings.append(
            "Recomputed net income differs from the sheet for: " + ", ".join(sorted(mismatches))
        )

    return entries


def parse_costs_sheet(sheet: Any, report: ImportReport) -> list[CostPosition]:
    """Parse ``Breakdown Kosten`` into cost positions."""
    headers, comment_col = _header_map(sheet)
    if not headers:
        report.warnings.append("Cost sheet: no owner columns found in the header row")
        return []

    # Fall back to the column right of the last owner column when the sheet
    # has comments but no header for them.
    if comment_col is None:
        comment_col = max(headers) + 1
    positions: list[CostPosition] = []
    seen_ids: set[str] = set()

    for row_idx in range(2, min(sheet.max_row, 200) + 1):
        name_value = sheet.cell(row=row_idx, column=1).value
        if name_value is None:
            continue
        name = str(name_value).strip()
        if not name:
            continue
        lowered = name.lower()
        # Skip the sheet's own summary rows — they are formulas over the list.
        if lowered in ("saldo", "total", "summe", "gesamt", "ausgaben"):
            continue

        comment = sheet.cell(row=row_idx, column=comment_col).value
        valid_from, valid_until = parse_validity(comment)
        is_buffer = any(marker in lowered for marker in _BUFFER_MARKERS)
        category = categorize_position(name)

        for col_idx, owner in headers.items():
            cell = sheet.cell(row=row_idx, column=col_idx)
            amount = _to_float(cell.value)
            if amount is None or amount == 0:
                continue

            buffer_units = None
            buffer_price = None
            kind = POSITION_KIND_FIXED
            if is_buffer:
                kind = POSITION_KIND_BUFFER
                buffer_units, buffer_price = _extract_buffer_factors(
                    sheet, row_idx, col_idx, amount
                )

            position_id = slugify_position(name, owner)
            if position_id in seen_ids:
                position_id = f"{position_id}_{row_idx}"
            seen_ids.add(position_id)

            positions.append(
                CostPosition(
                    id=position_id,
                    name=name,
                    owner=owner,
                    amount=round(amount, 2),
                    kind=kind,
                    category=category,
                    buffer_units=buffer_units,
                    buffer_unit_price=buffer_price,
                    valid_from=valid_from,
                    valid_until=valid_until,
                    note=str(comment).strip() if comment else "",
                )
            )

            if owner == OWNER_SHARED:
                report.shared_positions += 1
            if kind == POSITION_KIND_BUFFER:
                report.buffer_positions += 1
            if amount < 0:
                report.reimbursements += 1
            if valid_from or valid_until:
                report.time_boxed += 1

    report.positions_imported = len(positions)
    return positions


def _extract_buffer_factors(
    sheet: Any,
    row_idx: int,
    col_idx: int,
    amount: float,
) -> tuple[float | None, float | None]:
    """Recover ``(units, unit_price)`` for a buffer position.

    The spreadsheet writes buffers as a product (``=80*4.5``). Which factor is
    the count and which the price is not encoded, so the smaller factor is
    taken as the number of units — a weekly count is always smaller than a
    weekly price in this ledger. Both factors are editable afterwards, and an
    unparseable formula falls back to "1 unit at the full amount", which keeps
    the total correct.
    """
    raw = sheet.cell(row=row_idx, column=col_idx).value
    if isinstance(raw, str):
        match = _MULTIPLY_FORMULA.match(raw.strip())
        if match:
            a = float(match.group(1).replace(",", "."))
            b = float(match.group(2).replace(",", "."))
            units, price = (a, b) if a <= b else (b, a)
            return (units, price)
    return (1.0, round(amount, 2))


def parse_workbook(path: str) -> tuple[BudgetPlan, ImportReport]:
    """Parse a household workbook into a budget plan.

    Args:
        path: Absolute path to the ``.xlsx`` file.

    Returns:
        The parsed plan and a structural report.

    Raises:
        ImportError: openpyxl is not installed.
        FileNotFoundError: the path does not exist.
        ValueError: the workbook has neither expected sheet.
    """
    try:
        import openpyxl
    except ImportError as err:  # pragma: no cover - dependency declared in manifest
        raise ImportError(
            "openpyxl is required to import a spreadsheet. It is declared in the "
            "integration manifest; restart Home Assistant if it was just added."
        ) from err

    report = ImportReport()

    # Formulas are needed to recover buffer factors, cached values for amounts.
    formulas = openpyxl.load_workbook(path, data_only=False)
    values = openpyxl.load_workbook(path, data_only=True)

    income_sheet = _find_sheet(values, SHEET_INCOME)
    costs_values = _find_sheet(values, SHEET_COSTS)
    costs_formulas = _find_sheet(formulas, SHEET_COSTS)

    if income_sheet is None and costs_values is None:
        raise ValueError(
            "Workbook contains neither a 'Breakdown Einkommen' nor a "
            "'Breakdown Kosten' sheet — is this the household calculation?"
        )

    plan = BudgetPlan(source="spreadsheet-import")

    if income_sheet is not None:
        for entry in parse_income_sheet(income_sheet, report).values():
            plan.set_income(entry)
    else:
        report.warnings.append("No income sheet found — income must be entered manually")

    if costs_values is not None:
        positions = parse_costs_sheet(costs_values, report)
        # Re-read buffer factors from the formula view, which the value view lost.
        if costs_formulas is not None:
            _restore_buffer_factors(positions, costs_formulas)
        for position in positions:
            plan.upsert_position(position)
    else:
        report.warnings.append("No cost sheet found — positions must be entered manually")

    report.persons = list(plan.persons)
    _LOGGER.info(
        "Spreadsheet import: %d positions (%d shared, %d buffer, %d credits, %d time-boxed), "
        "%d persons — net income check: %s",
        report.positions_imported,
        report.shared_positions,
        report.buffer_positions,
        report.reimbursements,
        report.time_boxed,
        len(report.persons),
        report.net_income_check or "n/a",
    )
    return plan, report


def _restore_buffer_factors(positions: list[CostPosition], formula_sheet: Any) -> None:
    """Fill buffer factors from the formula view of the cost sheet.

    The cached-value workbook has already collapsed ``=80*4.5`` to ``360``, so
    the factors are looked up again by matching position names.
    """
    by_name: dict[str, str] = {}
    for row_idx in range(2, min(formula_sheet.max_row, 200) + 1):
        name = formula_sheet.cell(row=row_idx, column=1).value
        if name is None:
            continue
        for col_idx in range(2, 8):
            raw = formula_sheet.cell(row=row_idx, column=col_idx).value
            if isinstance(raw, str) and raw.startswith("="):
                by_name[str(name).strip().lower()] = raw
                break

    for position in positions:
        if position.kind != POSITION_KIND_BUFFER:
            continue
        raw = by_name.get(position.name.strip().lower())
        if not raw:
            continue
        match = _MULTIPLY_FORMULA.match(raw.strip())
        if not match:
            continue
        a = float(match.group(1).replace(",", "."))
        b = float(match.group(2).replace(",", "."))
        position.buffer_units, position.buffer_unit_price = (a, b) if a <= b else (b, a)
